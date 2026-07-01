"""EvaOne.AI Backend - AI Executive Operating System"""
import io
import os
import uuid
import logging
import asyncio
from pathlib import Path
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Literal

import requests
from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Form, Header, Query, Response, Cookie, Request, Depends
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field

from emergentintegrations.llm.chat import LlmChat, UserMessage
from emergentintegrations.llm.openai import OpenAISpeechToText, OpenAITextToSpeech

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger("evaone")

# ====================================================================
# RATE LIMITING (in-memory sliding window per user_id)
# Free/guest: 30 req/min. Paid: 300 req/min. Bypassed for owner/exec.
# ====================================================================
import time as _time
from collections import deque, defaultdict
_rate_windows: dict = defaultdict(deque)
_RATE_LIMITS = {
    "guest":   {"window": 60, "max": 15},
    "free":    {"window": 60, "max": 30},
    "creator": {"window": 60, "max": 120},
    "founder": {"window": 60, "max": 240},
    "executive": {"window": 60, "max": 480},
    "studio":  {"window": 60, "max": 900},
}
def _rate_key(user) -> tuple:
    if user.is_guest:
        return ("guest",)
    return (user.plan or "free",)

def check_rate_limit(user) -> None:
    if user.role in ("owner", "executive"):
        return
    key = ("guest" if user.is_guest else (user.plan or "free"))
    cfg = _RATE_LIMITS.get(key) or _RATE_LIMITS["free"]
    now = _time.time()
    window = _rate_windows[(user.user_id, key)]
    # Drop old
    while window and window[0] < now - cfg["window"]:
        window.popleft()
    if len(window) >= cfg["max"]:
        from fastapi import HTTPException as _HTTP
        raise _HTTP(status_code=429, detail=f"Rate limit hit — max {cfg['max']} requests / {cfg['window']}s. Slow down.")
    window.append(now)


# ====================================================================
# AUDIT LOGGING — write structured events for self-healing observability
# ====================================================================
async def audit_log(event: str, meta: dict | None = None, level: str = "info", user_id: str | None = None):
    try:
        # Note: db and datetime imported below in main body — this call is deferred until they exist.
        from datetime import datetime as _dt, timezone as _tz
        doc = {
            "id": f"audit_{uuid.uuid4().hex[:12]}",
            "event": event,
            "level": level,
            "user_id": user_id,
            "meta": meta or {},
            "at": _dt.now(_tz.utc).isoformat(),
        }
        # db is created later; use lazy import via globals()
        _db = globals().get("db")
        if _db is not None:
            await _db.system_audit.insert_one(doc)
    except Exception as e:
        logger.warning(f"audit_log failed: {e}")

# Import uuid so audit_log works at import-time (moved up)


# ====================================================================
# LLM RETRY WRAPPER — exponential backoff for transient errors
# ====================================================================
async def llm_send_with_retry(chat, user_message, *, max_retries: int = 2, base_delay: float = 0.6):
    import asyncio as _asyncio
    last_err = None
    for attempt in range(max_retries + 1):
        try:
            return await chat.send_message(user_message)
        except Exception as e:
            last_err = e
            msg = str(e).lower()
            transient = any(t in msg for t in ["rate", "timeout", "connection", "temporar", "overload", "503", "504", "502", "unavailable"])
            if not transient or attempt == max_retries:
                await audit_log("llm_call_failed", {"attempt": attempt + 1, "error": str(e)[:400]}, level="error")
                raise
            await audit_log("llm_call_retry", {"attempt": attempt + 1, "error": str(e)[:200]}, level="warn")
            await _asyncio.sleep(base_delay * (2 ** attempt))
    raise last_err  # unreachable



# ---------- MongoDB ----------
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# ---------- Constants ----------
EMERGENT_LLM_KEY = os.environ.get("EMERGENT_LLM_KEY")
APP_NAME = os.environ.get("APP_NAME", "evaone")
STORAGE_URL = "https://integrations.emergentagent.com/objstore/api/v1/storage"
AUTH_SESSION_URL = "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data"

# Available models
AVAILABLE_MODELS = {
    "claude-sonnet-4-6": {"provider": "anthropic", "label": "Claude Sonnet 4.6"},
    "gpt-5": {"provider": "openai", "label": "GPT-5"},
    "gemini-3.1-pro-preview": {"provider": "gemini", "label": "Gemini 3.1 Pro"},
    "gemini-3-flash-preview": {"provider": "gemini", "label": "Gemini 3 Flash"},
}
DEFAULT_MODEL = "claude-sonnet-4-6"

EVA_SYSTEM_PROMPT = (
    "You are Eva — the AI Chief of Staff of EvaOne.AI, an executive operating system. "
    "You are calm, sharp, decisive, and act as a senior executive partner — never as a chatbot. "
    "You help the user transform information into action: structure ideas, propose plans, summarize files, "
    "extract action items, and surface decisions. Respond with executive clarity: short headers, crisp bullets, "
    "and recommended next steps. When confidence is low, say so. Never claim to have performed external actions "
    "(emails, payments, deployments, posts) — only propose drafts for approval. Use markdown for structure.\n\n"
    "WHAT YOU CAN DO:\n"
    "- Chat, reason, plan, summarize, extract action items\n"
    "- Analyze uploaded files (PDF, DOCX, XLSX, CSV, TXT, images, audio)\n"
    "- Read & write to the user's Knowledge Vault (notes, memories)\n"
    "- Operate as voice (Whisper STT + OpenAI TTS) when the user enables it\n"
    "- Surface dashboards, projects, and open action items\n\n"
    "WHAT YOU CANNOT DO (BE TRANSPARENT — STATE THESE WHEN ASKED OR WHEN RELEVANT):\n"
    "- No access to your own configuration, subscription settings, or backend systems\n"
    "- No live internet, browsing, real-time pricing, or external lookups\n"
    "- No outbound actions yet: no sending emails, posting messages, executing payments, "
    "  triggering deployments, creating calendar events, or modifying third-party tools "
    "  (Gmail, Slack, Notion, HubSpot, Linear, etc.). Those integrations are Phase 2.\n"
    "- No memory of conversations you do not explicitly save to the Vault — "
    "  context retention is per-session unless the user pins it to the Vault.\n"
    "- No knowledge of events after your model's training cutoff unless the user provides it.\n\n"
    "When the user asks you to do something outside these bounds, say so plainly, "
    "then offer the closest thing you CAN do (e.g., 'I can't send the email, "
    "but I can draft it and add it to your Approval Queue once we ship that module')."
)

# ---------- Object Storage ----------
storage_key: Optional[str] = None

def init_storage():
    global storage_key
    if storage_key:
        return storage_key
    resp = requests.post(f"{STORAGE_URL}/init", json={"emergent_key": EMERGENT_LLM_KEY}, timeout=30)
    resp.raise_for_status()
    storage_key = resp.json()["storage_key"]
    return storage_key

def put_object(path: str, data: bytes, content_type: str) -> dict:
    key = init_storage()
    resp = requests.put(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key, "Content-Type": content_type},
        data=data, timeout=120
    )
    resp.raise_for_status()
    return resp.json()

def get_object(path: str) -> tuple[bytes, str]:
    key = init_storage()
    resp = requests.get(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key}, timeout=60
    )
    resp.raise_for_status()
    return resp.content, resp.headers.get("Content-Type", "application/octet-stream")

# ---------- FastAPI app ----------
app = FastAPI(title="EvaOne.AI")
api_router = APIRouter(prefix="/api")

@app.on_event("startup")
async def startup():
    try:
        init_storage()
        logger.info("Object storage initialized")
    except Exception as e:
        logger.error(f"Storage init failed: {e}")
    # Seed default workspace data
    await seed_demo_data()

@app.on_event("shutdown")
async def shutdown():
    client.close()

# ---------- Models ----------
ROLES = ["owner", "executive", "admin", "studio_operator", "member", "guest"]
PLANS = {
    "free":      {"id": "free",      "label": "Free",      "price": 0,    "chat_quota": 25,  "file_quota": 3,  "memory_quota": 10,  "models": ["claude-sonnet-4-6"], "features": []},
    "creator":   {"id": "creator",   "label": "Creator",   "price": 19,   "chat_quota": 250, "file_quota": 25, "memory_quota": 100, "models": ["claude-sonnet-4-6", "gpt-5", "gemini-3-flash-preview"], "features": ["vault", "long_term_memory", "multi_model"]},
    "founder":   {"id": "founder",   "label": "Founder",   "price": 49,   "chat_quota": 1000,"file_quota": 100,"memory_quota": 500, "models": ["claude-sonnet-4-6", "gpt-5", "gemini-3.1-pro-preview", "gemini-3-flash-preview"], "features": ["vault", "long_term_memory", "multi_model", "boardroom", "voice", "integrations", "agent_workflows"]},
    "executive": {"id": "executive", "label": "Executive", "price": 99,   "chat_quota": 5000,"file_quota": 500,"memory_quota": 2000,"models": ["claude-sonnet-4-6", "gpt-5", "gemini-3.1-pro-preview", "gemini-3-flash-preview"], "features": ["vault", "long_term_memory", "multi_model", "boardroom", "voice", "integrations", "agent_workflows", "advanced_automations", "priority_models", "team_collab"]},
    "studio":    {"id": "studio",    "label": "Studio",    "price": 299,  "chat_quota": 25000,"file_quota": 5000,"memory_quota": 10000,"models": ["claude-sonnet-4-6", "gpt-5", "gemini-3.1-pro-preview", "gemini-3-flash-preview"], "features": ["vault", "long_term_memory", "multi_model", "boardroom", "voice", "integrations", "agent_workflows", "advanced_automations", "priority_models", "team_collab", "multi_user", "shared_vault", "team_agents", "admin_controls", "analytics"]},
}
GUEST_QUOTAS = {"chat_quota": 5, "file_quota": 3, "memory_quota": 0}

# Role helpers
ROLE_ORDER = {"guest": 0, "member": 1, "studio_operator": 2, "admin": 3, "executive": 4, "owner": 5}
def role_at_least(role: str, minimum: str) -> bool:
    return ROLE_ORDER.get(role, -1) >= ROLE_ORDER.get(minimum, 99)

class User(BaseModel):
    user_id: str
    email: str
    name: str
    picture: Optional[str] = None
    created_at: str
    role: str = "member"
    plan: str = "free"
    is_guest: bool = False
    chat_used: int = 0
    file_used: int = 0
    plan_renews_at: Optional[str] = None

class ChatSession(BaseModel):
    id: str
    user_id: str
    title: str
    model: str
    created_at: str
    updated_at: str

class ChatMessage(BaseModel):
    id: str
    session_id: str
    role: Literal["user", "assistant", "system"]
    content: str
    model: Optional[str] = None
    created_at: str

class VaultNote(BaseModel):
    id: str
    user_id: str
    title: str
    content: str
    tags: List[str] = []
    pinned: bool = False
    created_at: str
    updated_at: str

class Project(BaseModel):
    id: str
    user_id: str
    name: str
    description: str
    status: Literal["active", "paused", "completed"] = "active"
    priority: Literal["low", "medium", "high", "critical"] = "medium"
    progress: int = 0
    created_at: str

class FileRecord(BaseModel):
    id: str
    user_id: str
    filename: str
    storage_path: str
    content_type: str
    size: int
    summary: Optional[str] = None
    key_points: List[str] = []
    action_items: List[str] = []
    extracted_text_preview: Optional[str] = None
    status: Literal["uploaded", "analyzing", "analyzed", "failed"] = "uploaded"
    created_at: str

# ---------- Auth ----------
async def get_current_user(request: Request, authorization: Optional[str] = Header(None)) -> User:
    token = request.cookies.get("session_token")
    if not token and authorization and authorization.startswith("Bearer "):
        token = authorization.split(" ", 1)[1]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    session = await db.user_sessions.find_one({"session_token": token}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    expires_at = session["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=401, detail="Session expired")
    user_doc = await db.users.find_one({"user_id": session["user_id"]}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=401, detail="User not found")
    # Backfill defaults for older docs
    user_doc.setdefault("role", "member")
    user_doc.setdefault("plan", "free")
    user_doc.setdefault("is_guest", False)
    user_doc.setdefault("chat_used", 0)
    user_doc.setdefault("file_used", 0)
    return User(**user_doc)

async def require_role(user: User, minimum: str) -> None:
    if not role_at_least(user.role, minimum):
        raise HTTPException(status_code=403, detail=f"Requires role: {minimum}")

def user_plan_features(user: User) -> set:
    return set(PLANS.get(user.plan, PLANS["free"])["features"])

async def require_feature(user: User, feature: str) -> None:
    # Owners + Executive role bypass plan gating
    if user.role in ("owner", "executive"):
        return
    if feature not in user_plan_features(user):
        raise HTTPException(
            status_code=402,
            detail=f"Feature '{feature}' requires upgrade. Current plan: {user.plan}",
        )

async def check_quota(user: User, kind: str) -> None:
    """Enforce monthly quotas for chat / file uploads. Owners/Executive bypass."""
    if user.role in ("owner", "executive"):
        return
    if user.is_guest:
        limits = GUEST_QUOTAS
    else:
        limits = PLANS.get(user.plan, PLANS["free"])
    if kind == "chat":
        used = user.chat_used or 0
        cap = limits.get("chat_quota", 0)
        if used >= cap:
            raise HTTPException(status_code=402, detail=f"Chat quota reached ({cap}). Upgrade to continue.")
    elif kind == "file":
        used = user.file_used or 0
        cap = limits.get("file_quota", 0)
        if used >= cap:
            raise HTTPException(status_code=402, detail=f"File quota reached ({cap}). Upgrade to continue.")

async def bump_usage(user_id: str, kind: str) -> None:
    field = "chat_used" if kind == "chat" else "file_used"
    await db.users.update_one({"user_id": user_id}, {"$inc": {field: 1}})

class SessionRequest(BaseModel):
    session_id: str

@api_router.post("/auth/session")
async def create_session(payload: SessionRequest, response: Response):
    """Exchange Emergent session_id for our own session_token, store user, set cookie."""
    try:
        r = requests.get(
            AUTH_SESSION_URL,
            headers={"X-Session-ID": payload.session_id},
            timeout=15
        )
        if r.status_code != 200:
            raise HTTPException(status_code=401, detail="Invalid session_id from Emergent Auth")
        data = r.json()
    except requests.RequestException:
        raise HTTPException(status_code=502, detail="Auth provider unreachable")

    email = data["email"]
    existing = await db.users.find_one({"email": email}, {"_id": 0})

    # Check if a real (non-guest) owner exists anywhere in the system
    real_owner_count = await db.users.count_documents({
        "role": "owner", "is_guest": {"$ne": True}
    })

    if existing:
        user_id = existing["user_id"]
        updates = {"name": data["name"], "picture": data.get("picture")}
        # Auto-promote to owner if no real owner exists yet (idempotent claim)
        owner_email = (os.environ.get("OWNER_EMAIL") or "").strip().lower()
        if existing.get("role") != "owner":
            if owner_email and email.lower() == owner_email:
                updates["role"] = "owner"
            elif real_owner_count == 0 and not existing.get("is_guest"):
                updates["role"] = "owner"
        await db.users.update_one({"user_id": user_id}, {"$set": updates})
    else:
        # New user: role assignment priority: OWNER_EMAIL → no-owner-yet → invite → member
        owner_email = (os.environ.get("OWNER_EMAIL") or "").strip().lower()
        assigned_role = "member"
        if owner_email and email.lower() == owner_email:
            assigned_role = "owner"
        elif real_owner_count == 0:
            assigned_role = "owner"
        else:
            invite = await db.invites.find_one({
                "email": email.lower(),
                "status": "pending",
                "expires_at": {"$gt": datetime.now(timezone.utc).isoformat()},
            }, {"_id": 0})
            if invite:
                assigned_role = invite["role"]
                await db.invites.update_one(
                    {"token": invite["token"]},
                    {"$set": {"status": "accepted", "accepted_at": datetime.now(timezone.utc).isoformat()}}
                )

        user_id = f"user_{uuid.uuid4().hex[:12]}"
        await db.users.insert_one({
            "user_id": user_id,
            "email": email,
            "name": data["name"],
            "picture": data.get("picture"),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "role": assigned_role,
            "plan": "free",
            "is_guest": False,
            "chat_used": 0,
            "file_used": 0,
        })

    session_token = data["session_token"]
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    await db.user_sessions.insert_one({
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": expires_at,
        "created_at": datetime.now(timezone.utc)
    })

    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
        max_age=7 * 24 * 60 * 60,
    )
    user_doc = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    return {"user": user_doc, "session_token": session_token}

@api_router.get("/auth/me")
async def auth_me(user: User = Depends(get_current_user)):
    return user.model_dump()

@api_router.post("/auth/logout")
async def auth_logout(request: Request, response: Response):
    token = request.cookies.get("session_token")
    if token:
        await db.user_sessions.delete_one({"session_token": token})
    response.delete_cookie("session_token", path="/")
    return {"ok": True}


@api_router.post("/auth/claim-owner")
async def claim_owner(user: User = Depends(get_current_user)):
    """Self-service ownership claim. Only works when no real owner exists yet.
    Idempotent — if user is already owner or another real owner exists, no-op with clear response."""
    if user.is_guest:
        raise HTTPException(status_code=403, detail="Sign in with Google first")
    if user.role == "owner":
        return {"ok": True, "already_owner": True, "user": user.model_dump()}
    real_owner_count = await db.users.count_documents({
        "role": "owner", "is_guest": {"$ne": True}
    })
    if real_owner_count > 0:
        raise HTTPException(status_code=409, detail="An owner already exists for this workspace.")
    await db.users.update_one({"user_id": user.user_id}, {"$set": {"role": "owner"}})
    await audit_log("owner_claimed", {"user_id": user.user_id, "email": user.email})
    updated = await db.users.find_one({"user_id": user.user_id}, {"_id": 0})
    return {"ok": True, "already_owner": False, "user": updated}

@api_router.get("/auth/owner-status")
async def owner_status():
    """Public endpoint — reports whether an owner has been claimed yet."""
    count = await db.users.count_documents({"role": "owner", "is_guest": {"$ne": True}})
    return {"owner_claimed": count > 0}


# ---------- Models endpoint ----------
@api_router.get("/models")
async def list_models():
    return [
        {"id": k, "label": v["label"], "provider": v["provider"]}
        for k, v in AVAILABLE_MODELS.items()
    ]

# ---------- Chat ----------
class CreateSessionBody(BaseModel):
    title: Optional[str] = None
    model: Optional[str] = None

@api_router.post("/chat/sessions")
async def create_chat_session(body: CreateSessionBody, user: User = Depends(get_current_user)):
    session_id = f"sess_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": session_id,
        "user_id": user.user_id,
        "title": body.title or "New conversation",
        "model": body.model or DEFAULT_MODEL,
        "created_at": now,
        "updated_at": now,
    }
    await db.chat_sessions.insert_one(dict(doc))
    return doc

@api_router.get("/chat/sessions")
async def list_chat_sessions(user: User = Depends(get_current_user)):
    cur = db.chat_sessions.find({"user_id": user.user_id}, {"_id": 0}).sort("updated_at", -1)
    return await cur.to_list(200)

@api_router.delete("/chat/sessions/{session_id}")
async def delete_chat_session(session_id: str, user: User = Depends(get_current_user)):
    await db.chat_sessions.delete_one({"id": session_id, "user_id": user.user_id})
    await db.chat_messages.delete_many({"session_id": session_id})
    return {"ok": True}

@api_router.get("/chat/sessions/{session_id}/messages")
async def list_messages(session_id: str, user: User = Depends(get_current_user)):
    sess = await db.chat_sessions.find_one({"id": session_id, "user_id": user.user_id}, {"_id": 0})
    if not sess:
        raise HTTPException(status_code=404, detail="Session not found")
    cur = db.chat_messages.find({"session_id": session_id}, {"_id": 0}).sort("created_at", 1)
    return await cur.to_list(1000)

class SendMessageBody(BaseModel):
    content: str
    model: Optional[str] = None
    attach_vault_context: bool = False

@api_router.post("/chat/sessions/{session_id}/messages")
async def send_message(session_id: str, body: SendMessageBody, user: User = Depends(get_current_user)):
    sess = await db.chat_sessions.find_one({"id": session_id, "user_id": user.user_id}, {"_id": 0})
    if not sess:
        raise HTTPException(status_code=404, detail="Session not found")

    # Quota check for chats
    await check_quota(user, "chat")
    check_rate_limit(user)
    model_id = body.model or sess.get("model") or DEFAULT_MODEL
    if model_id not in AVAILABLE_MODELS:
        raise HTTPException(status_code=400, detail="Unknown model")
    # Model gating: non-default model requires multi_model feature
    if model_id != DEFAULT_MODEL:
        await require_feature(user, "multi_model")
    provider = AVAILABLE_MODELS[model_id]["provider"]

    now = datetime.now(timezone.utc).isoformat()
    user_msg = {
        "id": f"msg_{uuid.uuid4().hex[:12]}",
        "session_id": session_id,
        "role": "user",
        "content": body.content,
        "model": model_id,
        "created_at": now,
    }
    await db.chat_messages.insert_one(dict(user_msg))

    # Reconstruct history for stateless model: emergent LlmChat library is per-call
    history_cur = db.chat_messages.find({"session_id": session_id}, {"_id": 0}).sort("created_at", 1)
    history = await history_cur.to_list(200)

    # Build system prompt; include long-term memory + workspace state + optional vault context
    sys_prompt = EVA_SYSTEM_PROMPT
    memory_ctx = await build_memory_context(user.user_id, limit=15)
    if memory_ctx:
        sys_prompt = sys_prompt + memory_ctx
    try:
        env_ctx = await build_env_context(user)
        sys_prompt = sys_prompt + env_ctx
    except Exception:
        pass
    if body.attach_vault_context:
        notes_cur = db.vault_notes.find({"user_id": user.user_id}, {"_id": 0}).sort("updated_at", -1).limit(10)
        notes = await notes_cur.to_list(10)
        if notes:
            ctx = "\n\n---\nKNOWLEDGE VAULT CONTEXT (most recent notes):\n" + "\n".join(
                f"- {n['title']}: {n['content'][:400]}" for n in notes
            )
            sys_prompt = sys_prompt + ctx

    chat = LlmChat(
        api_key=EMERGENT_LLM_KEY,
        session_id=session_id,
        system_message=sys_prompt,
    ).with_model(provider, model_id)

    # Replay history through chat — emergentintegrations LlmChat handles history per session_id internally
    # but to be safe we send only the latest message; library + session_id retains context
    try:
        assistant_text = await llm_send_with_retry(chat, UserMessage(text=body.content))
    except Exception as e:
        logger.exception("LLM call failed")
        await audit_log("chat_send_failed", {"session_id": session_id, "error": str(e)[:400]}, level="error", user_id=user.user_id)
        raise HTTPException(status_code=503, detail="Eva is temporarily unavailable. Please try again in a moment.")

    now2 = datetime.now(timezone.utc).isoformat()
    assistant_msg = {
        "id": f"msg_{uuid.uuid4().hex[:12]}",
        "session_id": session_id,
        "role": "assistant",
        "content": str(assistant_text),
        "model": model_id,
        "created_at": now2,
    }
    await db.chat_messages.insert_one(dict(assistant_msg))
    await bump_usage(user.user_id, "chat")

    # Update session title if first user turn
    if len(history) <= 1:
        title = body.content[:50] + ("..." if len(body.content) > 50 else "")
        await db.chat_sessions.update_one(
            {"id": session_id},
            {"$set": {"title": title, "updated_at": now2, "model": model_id}}
        )
    else:
        await db.chat_sessions.update_one(
            {"id": session_id},
            {"$set": {"updated_at": now2, "model": model_id}}
        )

    return {"user_message": user_msg, "assistant_message": assistant_msg}

# ---------- Voice ----------
@api_router.post("/voice/transcribe")
async def voice_transcribe(audio: UploadFile = File(...), user: User = Depends(get_current_user)):
    if audio.size and audio.size > 25 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Audio file too large (25MB max)")
    data = await audio.read()
    stt = OpenAISpeechToText(api_key=EMERGENT_LLM_KEY)
    try:
        # Wrap as file-like object
        buf = io.BytesIO(data)
        buf.name = audio.filename or "audio.webm"
        response = await stt.transcribe(file=buf, model="whisper-1", response_format="json")
        text = getattr(response, "text", None) or (response.get("text") if isinstance(response, dict) else str(response))
        return {"text": text}
    except Exception as e:
        logger.exception("STT failed")
        raise HTTPException(status_code=500, detail=f"Transcription failed: {e}")

class TTSBody(BaseModel):
    text: str
    voice: str = "nova"

@api_router.post("/voice/speak")
async def voice_speak(body: TTSBody, user: User = Depends(get_current_user)):
    if not body.text.strip():
        raise HTTPException(status_code=400, detail="Empty text")
    text = body.text[:4000]
    tts = OpenAITextToSpeech(api_key=EMERGENT_LLM_KEY)
    try:
        audio_bytes = await tts.generate_speech(text=text, model="tts-1", voice=body.voice)
        return Response(content=audio_bytes, media_type="audio/mpeg")
    except Exception as e:
        logger.exception("TTS failed")
        raise HTTPException(status_code=500, detail=f"TTS failed: {e}")

# ---------- Files ----------
def extract_text_from_file(filename: str, content_type: str, data: bytes) -> str:
    """Best-effort text extraction. Returns '' if not extractable."""
    ext = (filename.rsplit(".", 1)[-1] if "." in filename else "").lower()
    try:
        if ext == "pdf" or content_type == "application/pdf":
            from PyPDF2 import PdfReader
            reader = PdfReader(io.BytesIO(data))
            parts = []
            for page in reader.pages[:50]:
                try:
                    parts.append(page.extract_text() or "")
                except Exception:
                    pass
            return "\n".join(parts)
        if ext == "docx":
            import docx
            d = docx.Document(io.BytesIO(data))
            return "\n".join(p.text for p in d.paragraphs)
        if ext in ("xlsx", "xls"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            out = []
            for ws in wb.worksheets[:5]:
                out.append(f"# Sheet: {ws.title}")
                for row in ws.iter_rows(values_only=True, max_row=500):
                    out.append(" | ".join("" if c is None else str(c) for c in row))
            return "\n".join(out)
        if ext in ("csv", "txt", "json", "md", "log"):
            return data.decode("utf-8", errors="ignore")
    except Exception as e:
        logger.warning(f"Text extraction failed for {filename}: {e}")
    return ""

@api_router.post("/files/upload")
async def upload_file(file: UploadFile = File(...), user: User = Depends(get_current_user)):
    await check_quota(user, "file")
    data = await file.read()
    if len(data) > 25 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (25MB max)")
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in (file.filename or "") else "bin"
    path = f"{APP_NAME}/uploads/{user.user_id}/{uuid.uuid4().hex}.{ext}"
    content_type = file.content_type or "application/octet-stream"
    try:
        result = put_object(path, data, content_type)
    except Exception as e:
        logger.exception("Storage upload failed")
        raise HTTPException(status_code=500, detail=f"Storage upload failed: {e}")

    extracted = extract_text_from_file(file.filename or "", content_type, data)
    record = {
        "id": f"file_{uuid.uuid4().hex[:12]}",
        "user_id": user.user_id,
        "filename": file.filename or "untitled",
        "storage_path": result["path"],
        "content_type": content_type,
        "size": result.get("size", len(data)),
        "summary": None,
        "key_points": [],
        "action_items": [],
        "extracted_text_preview": (extracted[:2000] if extracted else None),
        "_full_text": extracted[:60000] if extracted else None,
        "status": "uploaded",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "is_deleted": False,
    }
    await db.files.insert_one(dict(record))
    await bump_usage(user.user_id, "file")
    record.pop("_full_text", None)
    return record

@api_router.get("/files")
async def list_files(user: User = Depends(get_current_user)):
    cur = db.files.find(
        {"user_id": user.user_id, "is_deleted": False},
        {"_id": 0, "_full_text": 0}
    ).sort("created_at", -1)
    return await cur.to_list(500)

@api_router.get("/files/{file_id}")
async def get_file(file_id: str, user: User = Depends(get_current_user)):
    rec = await db.files.find_one(
        {"id": file_id, "user_id": user.user_id, "is_deleted": False},
        {"_id": 0, "_full_text": 0}
    )
    if not rec:
        raise HTTPException(status_code=404, detail="File not found")
    return rec

@api_router.delete("/files/{file_id}")
async def delete_file(file_id: str, user: User = Depends(get_current_user)):
    res = await db.files.update_one(
        {"id": file_id, "user_id": user.user_id},
        {"$set": {"is_deleted": True}}
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="File not found")
    return {"ok": True}

@api_router.get("/files/{file_id}/download")
async def download_file(file_id: str, user: User = Depends(get_current_user)):
    rec = await db.files.find_one(
        {"id": file_id, "user_id": user.user_id, "is_deleted": False},
        {"_id": 0}
    )
    if not rec:
        raise HTTPException(status_code=404, detail="File not found")
    data, ct = get_object(rec["storage_path"])
    return Response(content=data, media_type=rec.get("content_type") or ct)

class AnalyzeBody(BaseModel):
    model: Optional[str] = None

@api_router.post("/files/{file_id}/analyze")
async def analyze_file(file_id: str, body: AnalyzeBody, user: User = Depends(get_current_user)):
    rec = await db.files.find_one({"id": file_id, "user_id": user.user_id, "is_deleted": False})
    if not rec:
        raise HTTPException(status_code=404, detail="File not found")
    text = rec.get("_full_text") or rec.get("extracted_text_preview")
    if not text:
        raise HTTPException(status_code=400, detail="No extractable text for this file type")

    model_id = body.model or DEFAULT_MODEL
    if model_id not in AVAILABLE_MODELS:
        model_id = DEFAULT_MODEL
    provider = AVAILABLE_MODELS[model_id]["provider"]

    await db.files.update_one({"id": file_id}, {"$set": {"status": "analyzing"}})

    sys = (
        "You are Eva, an executive AI analyst. Given the document text, return a structured JSON "
        "with: summary (3-5 sentences), key_points (5-8 bullets), action_items (3-7 concrete next steps for the user). "
        "Return STRICT JSON only, no markdown fencing, no commentary. Schema: "
        '{"summary": str, "key_points": [str], "action_items": [str]}'
    )
    chat = LlmChat(
        api_key=EMERGENT_LLM_KEY,
        session_id=f"analyze_{file_id}",
        system_message=sys,
    ).with_model(provider, model_id)

    prompt = f"FILENAME: {rec['filename']}\n\nDOCUMENT:\n{text[:50000]}"
    try:
        raw = await llm_send_with_retry(chat, UserMessage(text=prompt))
    except Exception as e:
        await db.files.update_one({"id": file_id}, {"$set": {"status": "failed"}})
        await audit_log("file_analysis_failed", {"file_id": file_id, "error": str(e)[:400]}, level="error", user_id=user.user_id)
        raise HTTPException(status_code=503, detail="Eva analyzer is temporarily unavailable. Please try again.")

    import json
    import re
    raw_str = str(raw).strip()
    # Strip code fences if present
    raw_str = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw_str, flags=re.MULTILINE).strip()
    try:
        parsed = json.loads(raw_str)
    except Exception:
        # Try extracting JSON object
        m = re.search(r"\{[\s\S]*\}", raw_str)
        parsed = json.loads(m.group(0)) if m else {"summary": raw_str[:2000], "key_points": [], "action_items": []}

    summary = parsed.get("summary", "")
    key_points = parsed.get("key_points", []) or []
    action_items = parsed.get("action_items", []) or []
    await db.files.update_one(
        {"id": file_id},
        {"$set": {
            "summary": summary,
            "key_points": key_points,
            "action_items": action_items,
            "status": "analyzed"
        }}
    )
    out = await db.files.find_one({"id": file_id}, {"_id": 0, "_full_text": 0})
    return out

# ---------- Vault Notes ----------
class NoteCreate(BaseModel):
    title: str
    content: str
    tags: List[str] = []
    pinned: bool = False

class NoteUpdate(BaseModel):
    title: Optional[str] = None
    content: Optional[str] = None
    tags: Optional[List[str]] = None
    pinned: Optional[bool] = None

@api_router.post("/vault/notes")
async def create_note(body: NoteCreate, user: User = Depends(get_current_user)):
    now = datetime.now(timezone.utc).isoformat()
    note = {
        "id": f"note_{uuid.uuid4().hex[:12]}",
        "user_id": user.user_id,
        "title": body.title,
        "content": body.content,
        "tags": body.tags,
        "pinned": body.pinned,
        "created_at": now,
        "updated_at": now,
    }
    await db.vault_notes.insert_one(dict(note))
    return note

@api_router.get("/vault/notes")
async def list_notes(q: Optional[str] = None, tag: Optional[str] = None, user: User = Depends(get_current_user)):
    query = {"user_id": user.user_id}
    if tag:
        query["tags"] = tag
    if q:
        query["$or"] = [
            {"title": {"$regex": q, "$options": "i"}},
            {"content": {"$regex": q, "$options": "i"}},
        ]
    cur = db.vault_notes.find(query, {"_id": 0}).sort([("pinned", -1), ("updated_at", -1)])
    return await cur.to_list(500)

@api_router.get("/vault/notes/{note_id}")
async def get_note(note_id: str, user: User = Depends(get_current_user)):
    n = await db.vault_notes.find_one({"id": note_id, "user_id": user.user_id}, {"_id": 0})
    if not n:
        raise HTTPException(status_code=404, detail="Note not found")
    return n

@api_router.put("/vault/notes/{note_id}")
async def update_note(note_id: str, body: NoteUpdate, user: User = Depends(get_current_user)):
    updates = {k: v for k, v in body.model_dump(exclude_none=True).items()}
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    res = await db.vault_notes.update_one(
        {"id": note_id, "user_id": user.user_id},
        {"$set": updates}
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Note not found")
    return await db.vault_notes.find_one({"id": note_id}, {"_id": 0})

@api_router.delete("/vault/notes/{note_id}")
async def delete_note(note_id: str, user: User = Depends(get_current_user)):
    await db.vault_notes.delete_one({"id": note_id, "user_id": user.user_id})
    return {"ok": True}

@api_router.get("/vault/search")
async def vault_search(q: str, user: User = Depends(get_current_user)):
    notes_cur = db.vault_notes.find({
        "user_id": user.user_id,
        "$or": [
            {"title": {"$regex": q, "$options": "i"}},
            {"content": {"$regex": q, "$options": "i"}},
        ]
    }, {"_id": 0}).limit(50)
    files_cur = db.files.find({
        "user_id": user.user_id,
        "is_deleted": False,
        "$or": [
            {"filename": {"$regex": q, "$options": "i"}},
            {"summary": {"$regex": q, "$options": "i"}},
            {"extracted_text_preview": {"$regex": q, "$options": "i"}},
        ]
    }, {"_id": 0, "_full_text": 0}).limit(50)
    return {
        "notes": await notes_cur.to_list(50),
        "files": await files_cur.to_list(50),
    }

# ---------- Projects ----------
class ProjectCreate(BaseModel):
    name: str
    description: str = ""
    priority: Literal["low", "medium", "high", "critical"] = "medium"
    progress: int = 0
    status: Literal["active", "paused", "completed"] = "active"

class ProjectUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    priority: Optional[Literal["low", "medium", "high", "critical"]] = None
    progress: Optional[int] = None
    status: Optional[Literal["active", "paused", "completed"]] = None

@api_router.post("/projects")
async def create_project(body: ProjectCreate, user: User = Depends(get_current_user)):
    p = {
        "id": f"proj_{uuid.uuid4().hex[:12]}",
        "user_id": user.user_id,
        **body.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.projects.insert_one(dict(p))
    return p

@api_router.get("/projects")
async def list_projects(user: User = Depends(get_current_user)):
    cur = db.projects.find({"user_id": user.user_id}, {"_id": 0}).sort("created_at", -1)
    return await cur.to_list(200)

@api_router.put("/projects/{project_id}")
async def update_project(project_id: str, body: ProjectUpdate, user: User = Depends(get_current_user)):
    updates = {k: v for k, v in body.model_dump(exclude_none=True).items()}
    res = await db.projects.update_one(
        {"id": project_id, "user_id": user.user_id},
        {"$set": updates}
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    return await db.projects.find_one({"id": project_id}, {"_id": 0})

@api_router.delete("/projects/{project_id}")
async def delete_project(project_id: str, user: User = Depends(get_current_user)):
    await db.projects.delete_one({"id": project_id, "user_id": user.user_id})
    return {"ok": True}

# ---------- Dashboard ----------
@api_router.get("/dashboard/stats")
async def dashboard_stats(user: User = Depends(get_current_user)):
    sessions = await db.chat_sessions.count_documents({"user_id": user.user_id})
    # Bounded session-id lookup to avoid unbounded scan under many sessions
    if sessions:
        session_ids = [s["id"] async for s in db.chat_sessions.find(
            {"user_id": user.user_id}, {"id": 1}
        ).limit(1000)]
        messages = await db.chat_messages.count_documents({"session_id": {"$in": session_ids}})
    else:
        messages = 0
    files = await db.files.count_documents({"user_id": user.user_id, "is_deleted": False})
    notes = await db.vault_notes.count_documents({"user_id": user.user_id})
    projects_total = await db.projects.count_documents({"user_id": user.user_id})
    projects_active = await db.projects.count_documents({"user_id": user.user_id, "status": "active"})
    analyzed_files = await db.files.count_documents({"user_id": user.user_id, "status": "analyzed", "is_deleted": False})

    # Aggregate all action items across analyzed files (bounded)
    cur = db.files.find(
        {"user_id": user.user_id, "is_deleted": False, "status": "analyzed"},
        {"_id": 0, "action_items": 1, "filename": 1, "id": 1, "created_at": 1}
    ).limit(50)
    open_actions = []
    async for f in cur:
        for ai in (f.get("action_items") or [])[:5]:
            open_actions.append({"file_id": f["id"], "filename": f["filename"], "action": ai})

    return {
        "sessions": sessions,
        "messages": messages,
        "files": files,
        "analyzed_files": analyzed_files,
        "notes": notes,
        "projects_total": projects_total,
        "projects_active": projects_active,
        "boardroom_sessions": await db.boardroom_sessions.count_documents({"user_id": user.user_id}),
        "memories": await db.user_memory.count_documents({"user_id": user.user_id}),
        "pending_approvals": await db.approvals.count_documents({"user_id": user.user_id, "status": "pending"}),
        "open_actions": open_actions[:12],
        "system_health": {
            "llm": "operational",
            "storage": "operational" if storage_key else "degraded",
            "voice": "operational",
            "vault": "operational",
        },
    }

@api_router.get("/dashboard/activity")
async def dashboard_activity(user: User = Depends(get_current_user)):
    """Recent unified activity feed."""
    activity = []
    async for m in db.chat_messages.aggregate([
        {"$lookup": {"from": "chat_sessions", "localField": "session_id", "foreignField": "id", "as": "sess"}},
        {"$unwind": "$sess"},
        {"$match": {"sess.user_id": user.user_id}},
        {"$sort": {"created_at": -1}},
        {"$limit": 10},
        {"$project": {"_id": 0, "role": 1, "content": 1, "created_at": 1, "session_title": "$sess.title"}}
    ]):
        activity.append({
            "type": "chat",
            "title": f"{m['role'].title()} in '{m['session_title']}'",
            "preview": (m["content"][:120] + "...") if len(m["content"]) > 120 else m["content"],
            "at": m["created_at"],
        })
    async for f in db.files.find(
        {"user_id": user.user_id, "is_deleted": False},
        {"_id": 0}
    ).sort("created_at", -1).limit(10):
        activity.append({
            "type": "file",
            "title": f"File uploaded: {f['filename']}",
            "preview": (f.get("summary") or "Awaiting analysis")[:120],
            "at": f["created_at"],
        })
    async for n in db.vault_notes.find(
        {"user_id": user.user_id}, {"_id": 0}
    ).sort("updated_at", -1).limit(10):
        activity.append({
            "type": "note",
            "title": f"Note updated: {n['title']}",
            "preview": n["content"][:120],
            "at": n["updated_at"],
        })
    activity.sort(key=lambda a: a["at"], reverse=True)
    return activity[:20]

# ---------- Seed demo (no-op if exists) ----------
async def seed_demo_data():
    # Nothing to seed globally; per-user examples are created on first login
    pass

@api_router.post("/onboard")
async def onboard(user: User = Depends(get_current_user)):
    """Create starter notes + project the first time the user opens the app."""
    existing = await db.vault_notes.count_documents({"user_id": user.user_id})
    if existing > 0:
        return {"created": False}
    now = datetime.now(timezone.utc).isoformat()
    starter_notes = [
        {
            "id": f"note_{uuid.uuid4().hex[:12]}",
            "user_id": user.user_id,
            "title": "Welcome to EvaOne.AI",
            "content": "EvaOne is your AI Chief of Staff. Use the Vault to capture insights, upload files for analysis, and chat with Eva to turn information into action.",
            "tags": ["welcome", "guide"],
            "pinned": True,
            "created_at": now,
            "updated_at": now,
        },
        {
            "id": f"note_{uuid.uuid4().hex[:12]}",
            "user_id": user.user_id,
            "title": "How Eva thinks",
            "content": "Eva drafts, structures, and surfaces decisions — never claims external actions. All outbound work passes through an approval queue before execution.",
            "tags": ["principles"],
            "pinned": False,
            "created_at": now,
            "updated_at": now,
        },
    ]
    await db.vault_notes.insert_many(starter_notes)
    await db.projects.insert_one({
        "id": f"proj_{uuid.uuid4().hex[:12]}",
        "user_id": user.user_id,
        "name": "Launch EvaOne workspace",
        "description": "Set up vault, upload first file, and have a strategy session with Eva.",
        "status": "active",
        "priority": "high",
        "progress": 25,
        "created_at": now,
    })
    return {"created": True}


# ====================================================================
# LONG-TERM MEMORY MODULE
# Persistent user profile fragments Eva auto-injects into every chat.
# ====================================================================

MEMORY_CATEGORIES = ["working_style", "decision", "person", "priority", "preference", "context"]

class MemoryCreate(BaseModel):
    category: Literal["working_style", "decision", "person", "priority", "preference", "context"]
    label: str
    content: str
    importance: int = 5
    source: Optional[str] = None

class MemoryUpdate(BaseModel):
    label: Optional[str] = None
    content: Optional[str] = None
    importance: Optional[int] = None
    category: Optional[Literal["working_style", "decision", "person", "priority", "preference", "context"]] = None

@api_router.post("/memory")
async def create_memory(body: MemoryCreate, user: User = Depends(get_current_user)):
    now = datetime.now(timezone.utc).isoformat()
    m = {
        "id": f"mem_{uuid.uuid4().hex[:12]}",
        "user_id": user.user_id,
        "category": body.category,
        "label": body.label,
        "content": body.content,
        "importance": max(1, min(10, body.importance)),
        "source": body.source or "manual",
        "created_at": now,
        "updated_at": now,
    }
    await db.user_memory.insert_one(dict(m))
    return m

@api_router.get("/memory")
async def list_memory(category: Optional[str] = None, user: User = Depends(get_current_user)):
    query = {"user_id": user.user_id}
    if category:
        query["category"] = category
    cur = db.user_memory.find(query, {"_id": 0}).sort([("importance", -1), ("updated_at", -1)])
    return await cur.to_list(500)

@api_router.put("/memory/{mem_id}")
async def update_memory(mem_id: str, body: MemoryUpdate, user: User = Depends(get_current_user)):
    updates = {k: v for k, v in body.model_dump(exclude_none=True).items()}
    if "importance" in updates:
        updates["importance"] = max(1, min(10, updates["importance"]))
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    res = await db.user_memory.update_one(
        {"id": mem_id, "user_id": user.user_id}, {"$set": updates}
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Memory not found")
    return await db.user_memory.find_one({"id": mem_id}, {"_id": 0})

@api_router.delete("/memory/{mem_id}")
async def delete_memory(mem_id: str, user: User = Depends(get_current_user)):
    await db.user_memory.delete_one({"id": mem_id, "user_id": user.user_id})
    return {"ok": True}

async def build_memory_context(user_id: str, limit: int = 20) -> str:
    cur = db.user_memory.find({"user_id": user_id}, {"_id": 0}).sort(
        [("importance", -1), ("updated_at", -1)]
    ).limit(limit)
    items = await cur.to_list(limit)
    if not items:
        return ""
    by_cat = {}
    for it in items:
        by_cat.setdefault(it["category"], []).append(f"- {it['label']}: {it['content']}")
    titles = {
        "working_style": "WORKING STYLE",
        "decision": "PRIOR DECISIONS",
        "person": "KEY PEOPLE",
        "priority": "CURRENT PRIORITIES",
        "preference": "PREFERENCES",
        "context": "ONGOING CONTEXT",
    }
    blocks = []
    for cat, rows in by_cat.items():
        blocks.append(f"### {titles.get(cat, cat.upper())}\n" + "\n".join(rows))
    return "\n\n---\nLONG-TERM USER MEMORY (always consider before responding):\n" + "\n\n".join(blocks)


# ====================================================================
# AI BOARDROOM MODULE - Virtual C-Suite Executive Debate
# ====================================================================

BOARDROOM_PERSONAS = [
    {"id": "ceo", "name": "CEO", "title": "The Visionary & Moderator", "color": "#00F0FF",
     "tone": "Decisive, strategic, macro-focused.",
     "lens": "Overall business strategy, market positioning, opportunity cost, synthesis."},
    {"id": "cpo", "name": "CPO", "title": "The User Champion", "color": "#8A2BE2",
     "tone": "Empathetic, detail-oriented, UX-obsessed.",
     "lens": "Product-market fit, user journeys, feature prioritization, retention."},
    {"id": "cto", "name": "CTO", "title": "The Practical Architect", "color": "#3DDC97",
     "tone": "Analytical, pragmatic, security-conscious.",
     "lens": "Tech stack, scalability, infrastructure, technical debt, security."},
    {"id": "cmo", "name": "CMO", "title": "The Growth Engine", "color": "#FFB454",
     "tone": "Energetic, data-driven, creative.",
     "lens": "User acquisition, branding, GTM strategy, viral loops, CAC."},
    {"id": "cfo", "name": "CFO", "title": "The Reality Check", "color": "#FF4D6D",
     "tone": "Conservative, precise, ROI-focused.",
     "lens": "Monetization, unit economics (LTV/CAC), burn rate, downside scenarios."},
    {"id": "coo", "name": "COO", "title": "The Operator", "color": "#7DD3FC",
     "tone": "Methodical, execution-focused, throughput-obsessed.",
     "lens": "Operations, process design, capacity planning, supply chain, delivery quality."},
    {"id": "legal", "name": "Legal", "title": "The Risk Counsel", "color": "#CBD5E1",
     "tone": "Cautious, precise, scenario-thinking.",
     "lens": "Regulatory exposure, contracts, IP, privacy, liability, defensibility."},
    {"id": "investor", "name": "Investor", "title": "The Capital Partner", "color": "#F472B6",
     "tone": "Pattern-matching, return-focused, blunt.",
     "lens": "Capital efficiency, fundability, milestone framing, dilution, exit math."},
]

BOARDROOM_SYSTEM = """You are the EvaOne AI Boardroom — an 8-persona executive debate engine.

PERSONAS (use as needed; CEO must open & synthesize):
- CEO — Visionary & Moderator (decisive, strategic, macro)
- CPO — User Champion (empathetic, UX-obsessed)
- CTO — Practical Architect (analytical, pragmatic, security-conscious)
- CMO — Growth Engine (energetic, data-driven, creative)
- CFO — Reality Check (conservative, precise, ROI-focused)
- COO — Operator (methodical, execution-focused, throughput)
- LEGAL — Risk Counsel (cautious, precise, regulation-aware)
- INVESTOR — Capital Partner (pattern-matching, return-focused, blunt)

RULES:
1. Each persona MUST have a distinct voice and challenge at least one other persona.
2. Structured FRICTION is mandatory — surface blind spots, not consensus.
3. Include CEO + minimum 4 other executives. Add COO/LEGAL/INVESTOR when the topic calls for them (ops, regulation, fundraising).
4. CEO synthesizes ONLY after all other executives have spoken.
5. Final action plan must be 3-7 prioritized, concrete next steps.
6. Output STRICT JSON ONLY (no markdown fencing, no commentary):

{
  "agenda": "1-sentence framing",
  "rounds": [
    {"persona": "ceo", "type": "opening", "content": "..."},
    {"persona": "cpo|cto|cmo|cfo|coo|legal|investor", "type": "argument", "content": "..."},
    ...
    {"persona": "X", "type": "rebuttal", "content": "...optional, pushes back on another exec..."}
  ],
  "synthesis": "CEO's final unified verdict (3-5 sentences)",
  "action_plan": ["concrete next step 1", "..."],
  "risks": ["risk 1", "risk 2"],
  "confidence": "low | medium | high"
}

Each persona's content: 2-4 focused sentences. Be specific. Use the user's actual context and numbers when given."""

class BoardroomSessionCreate(BaseModel):
    topic: str
    context: Optional[str] = None
    model: Optional[str] = None

@api_router.post("/boardroom/sessions")
async def create_boardroom_session(body: BoardroomSessionCreate, user: User = Depends(get_current_user)):
    await require_feature(user, "boardroom")
    sid = f"board_{uuid.uuid4().hex[:12]}"
    doc = {
        "id": sid,
        "user_id": user.user_id,
        "topic": body.topic,
        "context": body.context or "",
        "model": body.model or DEFAULT_MODEL,
        "status": "pending",
        "result": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.boardroom_sessions.insert_one(dict(doc))
    return doc

@api_router.get("/boardroom/sessions")
async def list_boardroom_sessions(user: User = Depends(get_current_user)):
    cur = db.boardroom_sessions.find({"user_id": user.user_id}, {"_id": 0}).sort("created_at", -1)
    return await cur.to_list(200)

@api_router.get("/boardroom/sessions/{sid}")
async def get_boardroom_session(sid: str, user: User = Depends(get_current_user)):
    doc = await db.boardroom_sessions.find_one({"id": sid, "user_id": user.user_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Session not found")
    return doc

@api_router.get("/boardroom/personas")
async def get_personas():
    return BOARDROOM_PERSONAS

@api_router.post("/boardroom/sessions/{sid}/run")
async def run_boardroom(sid: str, user: User = Depends(get_current_user)):
    doc = await db.boardroom_sessions.find_one({"id": sid, "user_id": user.user_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Session not found")
    model_id = doc.get("model") or DEFAULT_MODEL
    if model_id not in AVAILABLE_MODELS:
        model_id = DEFAULT_MODEL
    provider = AVAILABLE_MODELS[model_id]["provider"]

    await db.boardroom_sessions.update_one({"id": sid}, {"$set": {"status": "running"}})

    memory_ctx = await build_memory_context(user.user_id, limit=10)
    sys_prompt = BOARDROOM_SYSTEM + (("\n\n" + memory_ctx) if memory_ctx else "")

    chat = LlmChat(
        api_key=EMERGENT_LLM_KEY,
        session_id=f"board_{sid}",
        system_message=sys_prompt,
    ).with_model(provider, model_id)

    user_payload = f"TOPIC: {doc['topic']}"
    if doc.get("context"):
        user_payload += f"\n\nCONTEXT:\n{doc['context']}"

    try:
        raw = await llm_send_with_retry(chat, UserMessage(text=user_payload))
    except Exception as e:
        await db.boardroom_sessions.update_one({"id": sid}, {"$set": {"status": "failed"}})
        await audit_log("boardroom_failed", {"sid": sid, "error": str(e)[:400]}, level="error", user_id=user.user_id)
        raise HTTPException(status_code=503, detail="The board room is temporarily unavailable. Please try again.")

    import json
    import re
    raw_str = str(raw).strip()
    raw_str = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw_str, flags=re.MULTILINE).strip()
    try:
        parsed = json.loads(raw_str)
    except Exception:
        m = re.search(r"\{[\s\S]*\}", raw_str)
        if m:
            try:
                parsed = json.loads(m.group(0))
            except Exception:
                parsed = {"agenda": doc["topic"], "rounds": [], "synthesis": raw_str, "action_plan": [], "risks": [], "confidence": "low"}
        else:
            parsed = {"agenda": doc["topic"], "rounds": [], "synthesis": raw_str, "action_plan": [], "risks": [], "confidence": "low"}

    await db.boardroom_sessions.update_one(
        {"id": sid},
        {"$set": {"status": "complete", "result": parsed,
                  "completed_at": datetime.now(timezone.utc).isoformat()}}
    )
    return await db.boardroom_sessions.find_one({"id": sid}, {"_id": 0})

@api_router.delete("/boardroom/sessions/{sid}")
async def delete_boardroom_session(sid: str, user: User = Depends(get_current_user)):
    await db.boardroom_sessions.delete_one({"id": sid, "user_id": user.user_id})
    return {"ok": True}

class PublishBody(BaseModel):
    publish: bool

@api_router.post("/boardroom/sessions/{sid}/publish")
async def publish_boardroom(sid: str, body: PublishBody, user: User = Depends(get_current_user)):
    doc = await db.boardroom_sessions.find_one({"id": sid, "user_id": user.user_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Session not found")
    await db.boardroom_sessions.update_one(
        {"id": sid},
        {"$set": {
            "public": body.publish,
            "published_at": datetime.now(timezone.utc).isoformat() if body.publish else None,
        }}
    )
    return await db.boardroom_sessions.find_one({"id": sid}, {"_id": 0})

@app.get("/api/public/boardroom")
async def public_boardroom_list():
    """Publicly visible board sessions across all owners (no auth)."""
    cur = db.boardroom_sessions.find(
        {"public": True, "status": "complete"},
        {"_id": 0, "user_id": 0}
    ).sort("published_at", -1).limit(50)
    return await cur.to_list(50)

@app.get("/api/public/boardroom/{sid}")
async def public_boardroom_detail(sid: str):
    doc = await db.boardroom_sessions.find_one(
        {"id": sid, "public": True, "status": "complete"},
        {"_id": 0, "user_id": 0}
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Not found or not published")
    return doc


# ====================================================================
# APPROVAL QUEUE
# ====================================================================

class ApprovalDraftCreate(BaseModel):
    provider: Literal["gmail", "calendar", "slack", "notion", "hubspot", "internal"]
    action: str
    title: str
    payload: dict
    notes: Optional[str] = None

async def _insert_approval(user_id: str, body: ApprovalDraftCreate) -> dict:
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": f"app_{uuid.uuid4().hex[:12]}",
        "user_id": user_id,
        "provider": body.provider,
        "action": body.action,
        "title": body.title,
        "payload": body.payload,
        "notes": body.notes,
        "status": "pending",
        "execution_result": None,
        "created_at": now,
        "updated_at": now,
    }
    await db.approvals.insert_one(dict(doc))
    return doc

@api_router.post("/approvals")
async def create_approval(body: ApprovalDraftCreate, user: User = Depends(get_current_user)):
    return await _insert_approval(user.user_id, body)

@api_router.get("/approvals")
async def list_approvals(status: Optional[str] = None, user: User = Depends(get_current_user)):
    query = {"user_id": user.user_id}
    if status:
        query["status"] = status
    cur = db.approvals.find(query, {"_id": 0}).sort("created_at", -1)
    return await cur.to_list(500)

@api_router.post("/approvals/{aid}/approve")
async def approve_action(aid: str, user: User = Depends(get_current_user)):
    doc = await db.approvals.find_one({"id": aid, "user_id": user.user_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Approval not found")
    if doc["status"] != "pending":
        raise HTTPException(status_code=400, detail=f"Already {doc['status']}")
    mocked_result = {
        "executed": True,
        "mocked": True,
        "provider": doc["provider"],
        "executed_at": datetime.now(timezone.utc).isoformat(),
        "note": f"MOCKED execution. {doc['provider'].upper()} integration is not yet wired to a real API. The draft was preserved for audit.",
    }
    await db.approvals.update_one(
        {"id": aid},
        {"$set": {"status": "executed", "execution_result": mocked_result,
                  "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return await db.approvals.find_one({"id": aid}, {"_id": 0})

@api_router.post("/approvals/{aid}/reject")
async def reject_action(aid: str, user: User = Depends(get_current_user)):
    doc = await db.approvals.find_one({"id": aid, "user_id": user.user_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Approval not found")
    if doc["status"] != "pending":
        raise HTTPException(status_code=400, detail=f"Already {doc['status']}")
    await db.approvals.update_one(
        {"id": aid},
        {"$set": {"status": "rejected", "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return await db.approvals.find_one({"id": aid}, {"_id": 0})

@api_router.delete("/approvals/{aid}")
async def delete_approval(aid: str, user: User = Depends(get_current_user)):
    await db.approvals.delete_one({"id": aid, "user_id": user.user_id})
    return {"ok": True}


# ====================================================================
# INTEGRATION STUBS (MOCKED — create approval-queue drafts)
# ====================================================================

class GmailDraftBody(BaseModel):
    to: str
    subject: str
    body: str
    cc: Optional[str] = None

@api_router.post("/integrations/gmail/draft")
async def gmail_draft(body: GmailDraftBody, user: User = Depends(get_current_user)):
    return await _insert_approval(user.user_id, ApprovalDraftCreate(
        provider="gmail", action="send_email",
        title=f"Email → {body.to}: {body.subject}",
        payload=body.model_dump(),
        notes="MOCKED — requires GOOGLE_CLIENT_ID / GOOGLE_CLIENT_SECRET to execute.",
    ))

class CalendarDraftBody(BaseModel):
    title: str
    start: str
    end: str
    attendees: List[str] = []
    description: Optional[str] = ""

@api_router.post("/integrations/calendar/draft")
async def calendar_draft(body: CalendarDraftBody, user: User = Depends(get_current_user)):
    return await _insert_approval(user.user_id, ApprovalDraftCreate(
        provider="calendar", action="create_event",
        title=f"Event: {body.title}",
        payload=body.model_dump(),
        notes="MOCKED — requires Google OAuth to execute.",
    ))

class SlackDraftBody(BaseModel):
    channel: str
    message: str

@api_router.post("/integrations/slack/draft")
async def slack_draft(body: SlackDraftBody, user: User = Depends(get_current_user)):
    return await _insert_approval(user.user_id, ApprovalDraftCreate(
        provider="slack", action="post_message",
        title=f"Slack → #{body.channel}",
        payload=body.model_dump(),
        notes="MOCKED — requires SLACK_BOT_TOKEN to execute.",
    ))

class NotionDraftBody(BaseModel):
    title: str
    content: str
    parent_page: Optional[str] = None

@api_router.post("/integrations/notion/draft")
async def notion_draft(body: NotionDraftBody, user: User = Depends(get_current_user)):
    return await _insert_approval(user.user_id, ApprovalDraftCreate(
        provider="notion", action="create_page",
        title=f"Notion page: {body.title}",
        payload=body.model_dump(),
        notes="MOCKED — requires NOTION_TOKEN to execute.",
    ))

class HubspotDraftBody(BaseModel):
    object_type: Literal["contact", "deal", "company"]
    operation: Literal["create", "update"]
    fields: dict

@api_router.post("/integrations/hubspot/draft")
async def hubspot_draft(body: HubspotDraftBody, user: User = Depends(get_current_user)):
    return await _insert_approval(user.user_id, ApprovalDraftCreate(
        provider="hubspot", action=f"{body.operation}_{body.object_type}",
        title=f"HubSpot {body.operation} {body.object_type}",
        payload=body.model_dump(),
        notes="MOCKED — requires HUBSPOT_TOKEN to execute.",
    ))

@api_router.get("/integrations/status")
async def integrations_status():
    return {
        "gmail":    {"connected": False, "mocked": True, "env_var": "GOOGLE_CLIENT_ID"},
        "calendar": {"connected": False, "mocked": True, "env_var": "GOOGLE_CLIENT_ID"},
        "slack":    {"connected": bool(os.environ.get("SLACK_BOT_TOKEN")), "mocked": not bool(os.environ.get("SLACK_BOT_TOKEN")), "env_var": "SLACK_BOT_TOKEN"},
        "notion":   {"connected": bool(os.environ.get("NOTION_TOKEN")), "mocked": not bool(os.environ.get("NOTION_TOKEN")), "env_var": "NOTION_TOKEN"},
        "hubspot":  {"connected": bool(os.environ.get("HUBSPOT_TOKEN")), "mocked": not bool(os.environ.get("HUBSPOT_TOKEN")), "env_var": "HUBSPOT_TOKEN"},
    }



# ====================================================================
# PUBLIC + GUEST ACCESS
# ====================================================================

@api_router.post("/guest/start")
async def guest_start(response: Response):
    """Create an anonymous guest user with strict quotas. Returns session token."""
    user_id = f"guest_{uuid.uuid4().hex[:12]}"
    token = f"guest_{uuid.uuid4().hex}"
    now = datetime.now(timezone.utc)
    await db.users.insert_one({
        "user_id": user_id,
        "email": f"{user_id}@guest.evaone.local",
        "name": "Guest",
        "picture": None,
        "created_at": now.isoformat(),
        "role": "guest",
        "plan": "free",
        "is_guest": True,
        "chat_used": 0,
        "file_used": 0,
    })
    await db.user_sessions.insert_one({
        "user_id": user_id,
        "session_token": token,
        "expires_at": now + timedelta(days=1),
        "created_at": now,
    })
    response.set_cookie(
        key="session_token", value=token,
        httponly=True, secure=True, samesite="none",
        path="/", max_age=24 * 60 * 60,
    )
    user_doc = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    return {"user": user_doc, "session_token": token, "quotas": GUEST_QUOTAS}


# ====================================================================
# PLANS & BILLING (Stripe)
# ====================================================================

@api_router.get("/plans")
async def list_plans():
    return list(PLANS.values())

@api_router.get("/me/usage")
async def me_usage(user: User = Depends(get_current_user)):
    limits = GUEST_QUOTAS if user.is_guest else PLANS.get(user.plan, PLANS["free"])
    return {
        "plan": user.plan,
        "role": user.role,
        "is_guest": user.is_guest,
        "chat_used": user.chat_used or 0,
        "chat_quota": limits["chat_quota"],
        "file_used": user.file_used or 0,
        "file_quota": limits["file_quota"],
        "features": list(user_plan_features(user)) if not user.is_guest else [],
        "models": list(limits.get("models", [])),
    }

class CheckoutBody(BaseModel):
    plan_id: Literal["creator", "founder", "executive", "studio"]
    origin_url: str

@api_router.post("/billing/checkout")
async def billing_checkout(body: CheckoutBody, http_request: Request, user: User = Depends(get_current_user)):
    if user.is_guest:
        raise HTTPException(status_code=403, detail="Sign in first")
    plan = PLANS.get(body.plan_id)
    if not plan or plan["price"] <= 0:
        raise HTTPException(status_code=400, detail="Invalid plan")

    try:
        from emergentintegrations.payments.stripe.checkout import (
            StripeCheckout, CheckoutSessionRequest
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Stripe SDK unavailable: {e}")

    api_key = os.environ.get("STRIPE_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="STRIPE_API_KEY not configured")

    host_url = str(http_request.base_url).rstrip("/")
    webhook_url = f"{host_url}/api/webhook/stripe"
    sc = StripeCheckout(api_key=api_key, webhook_url=webhook_url)

    origin = body.origin_url.rstrip("/")
    success_url = f"{origin}/billing/success?session_id={{CHECKOUT_SESSION_ID}}"
    cancel_url = f"{origin}/pricing"

    req = CheckoutSessionRequest(
        amount=float(plan["price"]),
        currency="usd",
        success_url=success_url,
        cancel_url=cancel_url,
        metadata={
            "user_id": user.user_id,
            "user_email": user.email,
            "plan_id": body.plan_id,
        },
    )
    try:
        sess = await sc.create_checkout_session(req)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Stripe error: {e}")

    await db.payment_transactions.insert_one({
        "id": f"pay_{uuid.uuid4().hex[:12]}",
        "user_id": user.user_id,
        "user_email": user.email,
        "session_id": sess.session_id,
        "amount": float(plan["price"]),
        "currency": "usd",
        "plan_id": body.plan_id,
        "payment_status": "initiated",
        "status": "pending",
        "metadata": {"plan_id": body.plan_id, "user_id": user.user_id},
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    })
    return {"url": sess.url, "session_id": sess.session_id}

@api_router.get("/billing/status/{session_id}")
async def billing_status(session_id: str, http_request: Request, user: User = Depends(get_current_user)):
    try:
        from emergentintegrations.payments.stripe.checkout import StripeCheckout
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Stripe SDK unavailable: {e}")

    api_key = os.environ.get("STRIPE_API_KEY")
    host_url = str(http_request.base_url).rstrip("/")
    sc = StripeCheckout(api_key=api_key, webhook_url=f"{host_url}/api/webhook/stripe")
    try:
        status = await sc.get_checkout_status(session_id)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Stripe error: {e}")

    txn = await db.payment_transactions.find_one({"session_id": session_id, "user_id": user.user_id})
    if not txn:
        raise HTTPException(status_code=404, detail="Transaction not found")
    # Idempotent: only upgrade once
    already_processed = txn.get("status") in ("complete", "expired", "failed")
    new_status = "pending"
    if status.payment_status == "paid" and status.status == "complete":
        new_status = "complete"
    elif status.status == "expired":
        new_status = "expired"

    await db.payment_transactions.update_one(
        {"session_id": session_id},
        {"$set": {
            "payment_status": status.payment_status,
            "status": new_status,
            "stripe_status": status.status,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }}
    )
    if not already_processed and new_status == "complete":
        plan_id = txn.get("plan_id") or status.metadata.get("plan_id")
        if plan_id in PLANS:
            renew = (datetime.now(timezone.utc) + timedelta(days=30)).isoformat()
            await db.users.update_one(
                {"user_id": user.user_id},
                {"$set": {"plan": plan_id, "plan_renews_at": renew, "chat_used": 0, "file_used": 0}}
            )
    return {
        "payment_status": status.payment_status,
        "status": status.status,
        "amount_total": status.amount_total,
        "currency": status.currency,
        "plan_id": txn.get("plan_id"),
    }

@app.post("/api/webhook/stripe")
async def stripe_webhook(request: Request):
    try:
        from emergentintegrations.payments.stripe.checkout import StripeCheckout
    except Exception:
        return {"ok": False}
    api_key = os.environ.get("STRIPE_API_KEY")
    host_url = str(request.base_url).rstrip("/")
    sc = StripeCheckout(api_key=api_key, webhook_url=f"{host_url}/api/webhook/stripe")
    body = await request.body()
    sig = request.headers.get("Stripe-Signature", "")
    try:
        evt = await sc.handle_webhook(body, sig)
    except Exception as e:
        logger.warning(f"Stripe webhook verification failed: {e}")
        return {"ok": False}
    sid = evt.session_id
    if sid and evt.payment_status == "paid":
        txn = await db.payment_transactions.find_one({"session_id": sid})
        if txn and txn.get("status") != "complete":
            await db.payment_transactions.update_one(
                {"session_id": sid},
                {"$set": {"status": "complete", "payment_status": "paid",
                          "updated_at": datetime.now(timezone.utc).isoformat()}}
            )
            plan_id = txn.get("plan_id") or (evt.metadata or {}).get("plan_id")
            if plan_id in PLANS and txn.get("user_id"):
                renew = (datetime.now(timezone.utc) + timedelta(days=30)).isoformat()
                await db.users.update_one(
                    {"user_id": txn["user_id"]},
                    {"$set": {"plan": plan_id, "plan_renews_at": renew, "chat_used": 0, "file_used": 0}}
                )
    return {"ok": True}


# ====================================================================
# TEAM / INVITES
# ====================================================================

class InviteCreate(BaseModel):
    email: str
    role: Literal["executive", "admin", "studio_operator", "member"]
    expires_in_days: int = 7

@api_router.post("/invites")
async def create_invite(body: InviteCreate, user: User = Depends(get_current_user)):
    await require_role(user, "admin")
    token = uuid.uuid4().hex + uuid.uuid4().hex[:16]
    now = datetime.now(timezone.utc)
    doc = {
        "id": f"inv_{uuid.uuid4().hex[:12]}",
        "token": token,
        "email": body.email.lower(),
        "role": body.role,
        "invited_by": user.user_id,
        "invited_by_email": user.email,
        "status": "pending",
        "expires_at": (now + timedelta(days=max(1, min(60, body.expires_in_days)))).isoformat(),
        "created_at": now.isoformat(),
    }
    await db.invites.insert_one(dict(doc))
    return doc

@api_router.get("/invites")
async def list_invites(user: User = Depends(get_current_user)):
    await require_role(user, "admin")
    cur = db.invites.find({}, {"_id": 0}).sort("created_at", -1)
    return await cur.to_list(500)

@api_router.delete("/invites/{token}")
async def revoke_invite(token: str, user: User = Depends(get_current_user)):
    await require_role(user, "admin")
    await db.invites.delete_one({"token": token})
    return {"ok": True}

@api_router.get("/invites/preview/{token}")
async def preview_invite(token: str):
    inv = await db.invites.find_one({"token": token}, {"_id": 0})
    if not inv:
        raise HTTPException(status_code=404, detail="Invite not found")
    return {
        "email": inv["email"],
        "role": inv["role"],
        "status": inv["status"],
        "expires_at": inv["expires_at"],
        "invited_by_email": inv.get("invited_by_email"),
    }

@api_router.get("/team")
async def list_team(user: User = Depends(get_current_user)):
    await require_role(user, "admin")
    cur = db.users.find(
        {"is_guest": {"$ne": True}, "role": {"$in": ["owner", "executive", "admin", "studio_operator", "member"]}},
        {"_id": 0}
    ).sort("created_at", 1)
    return await cur.to_list(1000)

class RoleUpdate(BaseModel):
    role: Literal["executive", "admin", "studio_operator", "member"]

@api_router.put("/team/{user_id}/role")
async def update_team_role(user_id: str, body: RoleUpdate, user: User = Depends(get_current_user)):
    await require_role(user, "owner")
    target = await db.users.find_one({"user_id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    if target.get("role") == "owner":
        raise HTTPException(status_code=400, detail="Owner cannot be demoted")
    await db.users.update_one({"user_id": user_id}, {"$set": {"role": body.role}})
    return await db.users.find_one({"user_id": user_id}, {"_id": 0})


# ====================================================================
# ENVIRONMENT AWARENESS - Eva knows the workspace state
# ====================================================================

@api_router.get("/env/snapshot")
async def env_snapshot(user: User = Depends(get_current_user)):
    """Live snapshot of the user's workspace — used by Eva for context awareness."""
    files = await db.files.count_documents({"user_id": user.user_id, "is_deleted": False})
    notes = await db.vault_notes.count_documents({"user_id": user.user_id})
    projects = await db.projects.count_documents({"user_id": user.user_id, "status": "active"})
    memories = await db.user_memory.count_documents({"user_id": user.user_id})
    boardrooms = await db.boardroom_sessions.count_documents({"user_id": user.user_id})
    pending_approvals = await db.approvals.count_documents({"user_id": user.user_id, "status": "pending"})

    integ = {
        "gmail":    {"connected": False, "mocked": True},
        "calendar": {"connected": False, "mocked": True},
        "slack":    {"connected": bool(os.environ.get("SLACK_BOT_TOKEN")), "mocked": not bool(os.environ.get("SLACK_BOT_TOKEN"))},
        "notion":   {"connected": bool(os.environ.get("NOTION_TOKEN")), "mocked": not bool(os.environ.get("NOTION_TOKEN"))},
        "hubspot":  {"connected": bool(os.environ.get("HUBSPOT_TOKEN")), "mocked": not bool(os.environ.get("HUBSPOT_TOKEN"))},
    }

    return {
        "user": {"name": user.name, "role": user.role, "plan": user.plan, "is_guest": user.is_guest},
        "workspace": {
            "files": files,
            "notes": notes,
            "active_projects": projects,
            "memories": memories,
            "boardroom_sessions": boardrooms,
            "pending_approvals": pending_approvals,
        },
        "models_available": list(PLANS.get(user.plan, PLANS["free"]).get("models", [])) if not user.is_guest else ["claude-sonnet-4-6"],
        "integrations": integ,
        "features": list(user_plan_features(user)) if not user.is_guest else [],
    }

async def build_env_context(user: User) -> str:
    """Compile workspace state into a system prompt addendum for Eva."""
    snap = await env_snapshot(user)  # call our own helper
    ws = snap["workspace"]
    integ = snap["integrations"]
    connected = [k for k, v in integ.items() if v["connected"]]
    mocked = [k for k, v in integ.items() if v["mocked"]]
    feat = snap["features"]
    role = snap["user"]["role"]
    plan = snap["user"]["plan"]
    lines = [
        "\n\n---\nLIVE WORKSPACE STATE (use this to answer 'what can we work on' questions without asking the user):",
        f"- Role: {role} · Plan: {plan}",
        f"- Files: {ws['files']} · Vault notes: {ws['notes']} · Memories: {ws['memories']}",
        f"- Active projects: {ws['active_projects']} · Board meetings: {ws['boardroom_sessions']} · Pending approvals: {ws['pending_approvals']}",
        f"- Models available: {', '.join(snap['models_available'])}",
        f"- Integrations LIVE: {', '.join(connected) if connected else 'none yet'}",
        f"- Integrations MOCKED (drafts only): {', '.join(mocked) if mocked else 'none'}",
        f"- Plan features unlocked: {', '.join(feat) if feat else 'none — free tier'}",
    ]
    return "\n".join(lines)


# ====================================================================
# SYSTEM HEALTH CENTER (Responsible Self-Healing - monitoring only)
# ====================================================================

@api_router.get("/health/diagnostics")
async def health_diagnostics():
    """Read-only health snapshot of all subsystems + recent errors."""
    checks = []
    try:
        await db.command("ping")
        checks.append({"name": "MongoDB", "status": "operational"})
    except Exception as e:
        checks.append({"name": "MongoDB", "status": "degraded", "error": str(e)[:120]})
    checks.append({"name": "Object Storage", "status": "operational" if storage_key else "degraded"})
    checks.append({"name": "LLM Router", "status": "operational" if EMERGENT_LLM_KEY else "degraded"})
    checks.append({"name": "Voice Engine", "status": "operational" if EMERGENT_LLM_KEY else "degraded"})
    checks.append({"name": "Billing (Stripe)", "status": "operational" if os.environ.get("STRIPE_API_KEY") else "degraded"})

    # Recent errors from audit log (last 24h)
    since = (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()
    error_count = await db.system_audit.count_documents({"level": "error", "at": {"$gte": since}})
    warn_count = await db.system_audit.count_documents({"level": "warn", "at": {"$gte": since}})
    recent = await db.system_audit.find(
        {"level": {"$in": ["error", "warn"]}, "at": {"$gte": since}},
        {"_id": 0}
    ).sort("at", -1).limit(10).to_list(10)

    all_ok = all(c["status"] == "operational" for c in checks) and error_count == 0
    return {
        "overall": "operational" if all_ok else ("degraded" if error_count > 3 else "operational"),
        "checks": checks,
        "recent_errors_24h": error_count,
        "recent_warnings_24h": warn_count,
        "recent_events": recent,
        "self_healing_mode": "monitor_only",
        "features": ["auto_retry_llm", "audit_logging", "rate_limiting", "graceful_degradation"],
        "checked_at": datetime.now(timezone.utc).isoformat(),
    }

@api_router.get("/health/audit")
async def health_audit(limit: int = 50):
    cur = db.system_audit.find({}, {"_id": 0}).sort("at", -1).limit(min(limit, 500))
    return await cur.to_list(limit)


# ---------- Health ----------
@api_router.get("/")
async def root():
    return {"service": "EvaOne.AI", "status": "online", "version": "1.0"}

# ---------- Mount ----------
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)
